import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

# Add the project root to sys.path so we can import backend modules
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from backend.data_loader import load_data, get_available_currencies, get_current_data, get_future_data
from backend.strategies import run_simulation

# Configuration
DATA_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'Casestudy_Data.xlsx')
OUTPUT_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'FX_Hedging_Calculations.xlsx')
START_DATE_STR = '2024-07-01'
INVESTMENT_RETURN = 0.10
RISK_FREE_RATE = 0.05
TREE_STEPS = 100
USD_PER_CURRENCY = 250_000_000.0

def auto_adjust_columns(ws):
    """Adjusts column widths based on the maximum length of the content."""
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

def style_header(ws, row=1):
    """Applies styling and freezes the header row."""
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.freeze_panes = ws[f'A{row+1}']

def apply_currency_format(ws, cols, start_row, end_row):
    for col in cols:
        for r in range(start_row, end_row + 1):
            ws[f"{col}{r}"].number_format = '"$"#,##0.00'

def apply_percent_format(ws, cols, start_row, end_row):
    for col in cols:
        for r in range(start_row, end_row + 1):
            ws[f"{col}{r}"].number_format = '0.00%'

def main():
    print(f"Loading market data from {DATA_FILE}...")
    load_data(DATA_FILE)
    currencies = get_available_currencies()
    start_dt = datetime.strptime(START_DATE_STR, '%Y-%m-%d')
    
    print("Running Strategy A (No Hedge)...")
    res_A = run_simulation('A', start_dt, INVESTMENT_RETURN, RISK_FREE_RATE, TREE_STEPS, currencies, USD_PER_CURRENCY)
    print("Running Strategy B (Forward Hedge)...")
    res_B = run_simulation('B', start_dt, INVESTMENT_RETURN, RISK_FREE_RATE, TREE_STEPS, currencies, USD_PER_CURRENCY)
    print("Running Strategy C (Option Hedge)...")
    res_C = run_simulation('C', start_dt, INVESTMENT_RETURN, RISK_FREE_RATE, TREE_STEPS, currencies, USD_PER_CURRENCY)
    
    wb = Workbook()
    
    # ---------------------------------------------------------
    # Sheet 1: Inputs
    # ---------------------------------------------------------
    ws_inputs = wb.active
    ws_inputs.title = "Inputs & Market Data"
    
    # Simulation Params
    ws_inputs.append(["Parameter", "Value"])
    ws_inputs.append(["Total Investment (USD)", USD_PER_CURRENCY * len(currencies)])
    ws_inputs.append(["Allocation per Currency (USD)", USD_PER_CURRENCY])
    ws_inputs.append(["Investment Return", INVESTMENT_RETURN])
    ws_inputs.append(["Risk Free Rate", RISK_FREE_RATE])
    ws_inputs.append(["Time to Maturity (Years)", 1.0])
    ws_inputs.append(["Simulation Start Date", START_DATE_STR])
    ws_inputs.append([])
    
    ws_inputs['B2'].number_format = '"$"#,##0.00'
    ws_inputs['B3'].number_format = '"$"#,##0.00'
    ws_inputs['B4'].number_format = '0.00%'
    ws_inputs['B5'].number_format = '0.00%'
    
    style_header(ws_inputs, 1)
    
    # Market Data
    start_row = 9
    ws_inputs.append(["Currency Pair", "Spot (T=0)", "Forward (T=1Y)", "ATM Volatility", "Spot (T=1Y)"])
    for ccy in currencies:
        cur_data = get_current_data(start_dt, ccy)
        fut_data = get_future_data(start_dt, ccy)
        ws_inputs.append([
            ccy,
            cur_data['Spot'],
            cur_data.get('Forward', cur_data['Spot']),
            cur_data.get('ATM Vol', 0.1),
            fut_data['Spot']
        ])
    style_header(ws_inputs, start_row)
    apply_percent_format(ws_inputs, ['D'], start_row + 1, start_row + len(currencies))
    auto_adjust_columns(ws_inputs)

    # ---------------------------------------------------------
    # Sheet 2: Portfolio Setup
    # ---------------------------------------------------------
    ws_setup = wb.create_sheet(title="Portfolio Setup")
    headers = ["Currency", "USD Allocation", "Spot Start", "Spot End", "FX % Change", "Local Currency Purchased", "Investment Growth", "Local Currency at Maturity"]
    ws_setup.append(headers)
    
    for i, ccy in enumerate(currencies, start=2):
        cur_data = get_current_data(start_dt, ccy)
        fut_data = get_future_data(start_dt, ccy)
        spot_start = cur_data['Spot']
        spot_end = fut_data['Spot']
        fx_change = (spot_end / spot_start) - 1
        local_purchased = USD_PER_CURRENCY * spot_start
        growth = local_purchased * INVESTMENT_RETURN
        local_maturity = local_purchased + growth
        
        ws_setup.append([ccy, USD_PER_CURRENCY, spot_start, spot_end, fx_change, local_purchased, growth, local_maturity])
        
    style_header(ws_setup, 1)
    apply_currency_format(ws_setup, ['B'], 2, len(currencies) + 1)
    apply_percent_format(ws_setup, ['E'], 2, len(currencies) + 1)
    # Generic number format for local currency
    for col in ['F', 'G', 'H']:
        for r in range(2, len(currencies) + 2):
            ws_setup[f"{col}{r}"].number_format = '#,##0.00'
    auto_adjust_columns(ws_setup)

    # ---------------------------------------------------------
    # Sheet 3: Strategy A (No Hedge)
    # ---------------------------------------------------------
    ws_a = wb.create_sheet(title="Strategy A - No Hedge")
    ws_a.append(["Currency", "USD Invested", "Local Amount", "Investment Return", "Final Local Value", "Spot at Maturity", "USD Returned", "Profit (USD)", "Return %"])
    
    for i, ccy in enumerate(currencies, start=2):
        det = res_A['countries'][ccy]
        cur_data = get_current_data(start_dt, ccy)
        fut_data = get_future_data(start_dt, ccy)
        spot_start = cur_data['Spot']
        spot_end = fut_data['Spot']
        local_amt = USD_PER_CURRENCY * spot_start
        final_local = local_amt * (1 + INVESTMENT_RETURN)
        usd_returned = final_local / spot_end
        profit = usd_returned - USD_PER_CURRENCY
        ret_pct = profit / USD_PER_CURRENCY
        
        ws_a.append([ccy, USD_PER_CURRENCY, local_amt, INVESTMENT_RETURN, final_local, spot_end, usd_returned, profit, ret_pct])
        
    # Portfolio Totals
    tot_row = len(currencies) + 2
    ws_a.append(["PORTFOLIO TOTAL", USD_PER_CURRENCY * len(currencies), "", "", "", "", res_A['total_pnl'] + (USD_PER_CURRENCY * len(currencies)), res_A['total_pnl'], res_A['portfolio_return_pct'] / 100])
    
    style_header(ws_a, 1)
    apply_currency_format(ws_a, ['B', 'G', 'H'], 2, tot_row)
    apply_percent_format(ws_a, ['D', 'I'], 2, tot_row)
    for r in range(2, tot_row):
        ws_a[f"C{r}"].number_format = '#,##0.00'
        ws_a[f"E{r}"].number_format = '#,##0.00'
    auto_adjust_columns(ws_a)

    # ---------------------------------------------------------
    # Sheet 4: Strategy B (Forward Hedge)
    # ---------------------------------------------------------
    ws_b = wb.create_sheet(title="Strategy B - Forward Hedge")
    ws_b.append(["Currency", "Forward Rate", "Final Local Value", "USD Received", "Profit (USD)", "Return %"])
    
    for i, ccy in enumerate(currencies, start=2):
        cur_data = get_current_data(start_dt, ccy)
        spot_start = cur_data['Spot']
        fwd_rate = cur_data.get('Forward', spot_start)
        local_amt = USD_PER_CURRENCY * spot_start
        final_local = local_amt * (1 + INVESTMENT_RETURN)
        usd_received = final_local / fwd_rate
        profit = usd_received - USD_PER_CURRENCY
        ret_pct = profit / USD_PER_CURRENCY
        
        ws_b.append([ccy, fwd_rate, final_local, usd_received, profit, ret_pct])
        
    tot_row = len(currencies) + 2
    ws_b.append(["PORTFOLIO TOTAL", "", "", res_B['total_pnl'] + (USD_PER_CURRENCY * len(currencies)), res_B['total_pnl'], res_B['portfolio_return_pct'] / 100])
    
    style_header(ws_b, 1)
    apply_currency_format(ws_b, ['D', 'E'], 2, tot_row)
    apply_percent_format(ws_b, ['F'], 2, tot_row)
    for r in range(2, tot_row):
        ws_b[f"C{r}"].number_format = '#,##0.00'
    auto_adjust_columns(ws_b)

    # ---------------------------------------------------------
    # Sheet 5: Option Pricing
    # ---------------------------------------------------------
    ws_opt = wb.create_sheet(title="Option Pricing")
    ws_opt.append(["Currency", "Spot", "Strike", "Volatility", "Risk Free Rate", "Time", "CRR Premium (USD)", "Black Scholes Premium (USD)", "Difference (USD)", "Error %", "Delta", "Gamma", "Theta", "Vega", "Rho"])
    
    for i, ccy in enumerate(currencies, start=2):
        c_det = res_C['countries'][ccy]['details']
        cur_data = get_current_data(start_dt, ccy)
        greeks = c_det.get('greeks', {})
        diff = c_det['convergence_gap']
        err_pct = diff / c_det['premium_bs_usd'] if c_det['premium_bs_usd'] != 0 else 0
        
        ws_opt.append([
            ccy, cur_data['Spot'], cur_data['Spot'], cur_data.get('ATM Vol', 0.1), RISK_FREE_RATE, 1.0,
            c_det['premium_crr_usd'], c_det['premium_bs_usd'], diff, err_pct,
            greeks.get('delta', 0), greeks.get('gamma', 0), greeks.get('theta', 0), greeks.get('vega', 0), greeks.get('rho', 0)
        ])
        
    style_header(ws_opt, 1)
    apply_currency_format(ws_opt, ['G', 'H', 'I'], 2, len(currencies) + 1)
    apply_percent_format(ws_opt, ['D', 'E', 'J'], 2, len(currencies) + 1)
    auto_adjust_columns(ws_opt)

    # ---------------------------------------------------------
    # Sheet 6: Strategy C (Option Hedge)
    # ---------------------------------------------------------
    ws_c = wb.create_sheet(title="Strategy C - Option Hedge")
    ws_c.append(["Currency", "Investment Profit (USD)", "Premium Paid (USD)", "Option Payoff (USD)", "Net Profit (USD)", "Return %"])
    
    for i, ccy in enumerate(currencies, start=2):
        c_det = res_C['countries'][ccy]['details']
        c_stats = res_C['countries'][ccy]
        
        ws_c.append([
            ccy,
            c_det['investment_pnl'],
            -c_det['premium_usd'],
            c_det['payoff_usd'],
            c_stats['pnl'],
            c_stats['pnl'] / USD_PER_CURRENCY
        ])
        
    tot_row = len(currencies) + 2
    inv_profit_total = sum([res_C['countries'][c]['details']['investment_pnl'] for c in currencies])
    premium_total = sum([-res_C['countries'][c]['details']['premium_usd'] for c in currencies])
    payoff_total = sum([res_C['countries'][c]['details']['payoff_usd'] for c in currencies])
    
    ws_c.append(["PORTFOLIO TOTAL", inv_profit_total, premium_total, payoff_total, res_C['total_pnl'], res_C['portfolio_return_pct'] / 100])
    
    style_header(ws_c, 1)
    apply_currency_format(ws_c, ['B', 'C', 'D', 'E'], 2, tot_row)
    apply_percent_format(ws_c, ['F'], 2, tot_row)
    auto_adjust_columns(ws_c)

    # ---------------------------------------------------------
    # Sheet 7: Strategy Comparison
    # ---------------------------------------------------------
    ws_comp = wb.create_sheet(title="Strategy Comparison")
    ws_comp.append(["Strategy", "Total Final Value", "Total Profit", "Return %", "FX Risk", "Upside", "Cost", "Certainty", "Winner"])
    
    pnl_a = res_A['total_pnl']
    pnl_b = res_B['total_pnl']
    pnl_c = res_C['total_pnl']
    
    winner = "A" if pnl_a > pnl_b and pnl_a > pnl_c else ("B" if pnl_b > pnl_c else "C")
    
    tot_inv = USD_PER_CURRENCY * len(currencies)
    ws_comp.append(["Strategy A (No Hedge)", tot_inv + pnl_a, pnl_a, res_A['portfolio_return_pct'] / 100, "Full exposure", "Unlimited", "$0", "None", "Winner" if winner == "A" else ""])
    ws_comp.append(["Strategy B (Forward)", tot_inv + pnl_b, pnl_b, res_B['portfolio_return_pct'] / 100, "Zero", "Capped", "Implicit (Fwd disc.)", "Full", "Winner" if winner == "B" else ""])
    ws_comp.append(["Strategy C (Option)", tot_inv + pnl_c, pnl_c, res_C['portfolio_return_pct'] / 100, "Downside hedged", "Retained", "Premium Paid", "Partial", "Winner" if winner == "C" else ""])
    
    style_header(ws_comp, 1)
    apply_currency_format(ws_comp, ['B', 'C'], 2, 4)
    apply_percent_format(ws_comp, ['D'], 2, 4)
    auto_adjust_columns(ws_comp)
    
    # Generate Charts
    
    # 1. Strategy Returns Bar Chart
    chart_ret = BarChart()
    chart_ret.type = "col"
    chart_ret.style = 10
    chart_ret.title = "Strategy Return Comparison"
    chart_ret.y_axis.title = "Return %"
    chart_ret.x_axis.title = "Strategy"
    
    data_ret = Reference(ws_comp, min_col=4, min_row=1, max_row=4)
    cats_ret = Reference(ws_comp, min_col=1, min_row=2, max_row=4)
    chart_ret.add_data(data_ret, titles_from_data=True)
    chart_ret.set_categories(cats_ret)
    ws_comp.add_chart(chart_ret, "E7")
    
    # ---------------------------------------------------------
    # Sheet 8: Recommendation
    # ---------------------------------------------------------
    ws_rec = wb.create_sheet(title="Recommendation")
    ws_rec.append(["Category", "Summary"])
    
    ws_rec.append(["Highest Return", f"Strategy {winner} achieved the highest return for this specific historical period."])
    ws_rec.append(["Lowest Risk", "Strategy B (Forward Hedge) provides completely certain USD returns, eliminating all FX risk."])
    ws_rec.append(["Balanced Strategy", "Strategy C (Option Hedge) limits downside risk while preserving upside potential, at the cost of the option premium."])
    ws_rec.append(["Recommended Corporate Strategy", "Use Strategy B for 70% of exposure to lock in budget targets, and Strategy C for the remaining 30% to capture potential upside."])
    
    style_header(ws_rec, 1)
    ws_rec.column_dimensions['A'].width = 30
    ws_rec.column_dimensions['B'].width = 120
    
    # Generate Chart for Profit by Currency (Strategy A vs B vs C)
    # Re-use space on Recommendation sheet or put it in comparison.
    ws_chart_data = wb.create_sheet(title="Chart Data", index=len(wb.sheetnames))
    ws_chart_data.sheet_state = 'hidden' # Hide it as it's just for charting
    
    ws_chart_data.append(["Currency", "Strat A Profit", "Strat B Profit", "Strat C Profit"])
    for ccy in currencies:
        ws_chart_data.append([
            ccy,
            res_A['countries'][ccy]['pnl'],
            res_B['countries'][ccy]['pnl'],
            res_C['countries'][ccy]['pnl']
        ])
    
    chart_pnl = BarChart()
    chart_pnl.type = "col"
    chart_pnl.style = 10
    chart_pnl.grouping = "clustered"
    chart_pnl.title = "Profit by Currency"
    chart_pnl.y_axis.title = "Profit (USD)"
    chart_pnl.x_axis.title = "Currency"
    
    data_pnl = Reference(ws_chart_data, min_col=2, max_col=4, min_row=1, max_row=len(currencies)+1)
    cats_pnl = Reference(ws_chart_data, min_col=1, min_row=2, max_row=len(currencies)+1)
    chart_pnl.add_data(data_pnl, titles_from_data=True)
    chart_pnl.set_categories(cats_pnl)
    
    ws_rec.add_chart(chart_pnl, "A7")

    print(f"Saving Excel report to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    print("Export complete.")

if __name__ == "__main__":
    main()
